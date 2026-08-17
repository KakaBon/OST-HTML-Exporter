from __future__ import annotations
import sys, html, math, colorsys, warnings
from pathlib import Path
from datetime import datetime, date, time
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except Exception:
    tk = None

WANTED_SHEETS = ["OST Timeline", "OST Collection"]

# Some workbooks contain Excel data-validation extensions that openpyxl cannot preserve
# if the workbook were saved again. This exporter only reads the workbook and writes HTML,
# so the warning is harmless here and is hidden to keep the launcher output clean.
warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
    category=UserWarning,
)


# Excel/OpenXML indexed color table fallback.
# Avoid importing internal openpyxl constants because their names may differ by version.
INDEXED_COLORS = [
    "00000000","00FFFFFF","00FF0000","0000FF00","000000FF","00FFFF00","00FF00FF","0000FFFF","00000000","00FFFFFF",
    "00FF0000","0000FF00","000000FF","00FFFF00","00FF00FF","0000FFFF","00800000","00008000","00000080","00808000",
    "00800080","00008080","00C0C0C0","00808080","009999FF","00993366","00FFFFCC","00CCFFFF","00660066","00FF8080",
    "000066CC","00CCCCFF","00000080","00FF00FF","00FFFF00","0000FFFF","00800080","00800000","00008080","000000FF",
    "0000CCFF","00CCFFFF","00CCFFCC","00FFFF99","0099CCFF","00FF99CC","00CC99FF","00FFCC99","003366FF","0033CCCC",
    "0099CC00","00FFCC00","00FF9900","00FF6600","00666699","00969696","00003366","00339966","00003300","00333300",
    "00993300","00993366","00333399","00333333"
]

def pick_file():
    if tk is None:
        return None
    root = tk.Tk()
    root.withdraw()
    path = filedialog.askopenfilename(
        title="Select Excel workbook",
        filetypes=[
            ("Excel workbooks", "*.xlsm *.xlsx"),
            ("All files", "*.*"),
        ],
    )
    root.destroy()
    return path or None

def show_info(title, msg):
    if tk is not None:
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo(title, msg)
        root.destroy()
    else:
        print(msg)

def show_error(title, msg):
    if tk is not None:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror(title, msg)
        root.destroy()
    else:
        print(msg, file=sys.stderr)

def _theme_colors_from_workbook(wb):
    """
    Resolve workbook theme colors in the numeric order used by Excel/openpyxl:
    0 lt1, 1 dk1, 2 lt2, 3 dk2, 4..9 accent1..accent6,
    10 hlink, 11 folHlink.

    The XML clrScheme itself is stored as dk1, lt1, dk2, lt2, ...,
    so using raw child order swaps the first four theme indices and can turn
    a light neutral gray into a dark blue-gray.
    """
    theme = getattr(wb, "loaded_theme", None)
    if not theme:
        return []

    try:
        root = ET.fromstring(theme)
        ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
        scheme = root.find(".//a:themeElements/a:clrScheme", ns)
        if scheme is None:
            return []

        by_name = {}
        for entry in list(scheme):
            name = entry.tag.split("}")[-1]
            children = list(entry)
            if not children:
                by_name[name] = None
                continue

            color_node = children[0]
            rgb = color_node.attrib.get("val")
            if color_node.tag.endswith("sysClr"):
                rgb = color_node.attrib.get("lastClr") or rgb

            by_name[name] = rgb[-6:].upper() if rgb and len(rgb) >= 6 else None

        order = [
            "lt1", "dk1", "lt2", "dk2",
            "accent1", "accent2", "accent3", "accent4", "accent5", "accent6",
            "hlink", "folHlink",
        ]
        return [by_name.get(name) for name in order]
    except Exception:
        return []


def _apply_excel_tint(rgb_hex, tint):
    """
    Apply Excel/OpenXML tint using the HLSMAX=240 algorithm used by Excel-style
    theme color resolution.

    tint < 0 darkens the luminance.
    tint > 0 lightens the luminance toward HLSMAX.
    """
    if not rgb_hex:
        return None

    try:
        tint = float(tint or 0.0)
        if abs(tint) < 1e-12:
            return rgb_hex.upper()

        HLSMAX = 240.0

        r = int(rgb_hex[0:2], 16) / 255.0
        g = int(rgb_hex[2:4], 16) / 255.0
        b = int(rgb_hex[4:6], 16) / 255.0

        h, l, s = colorsys.rgb_to_hls(r, g, b)
        h *= HLSMAX
        l *= HLSMAX
        s *= HLSMAX

        if tint < 0:
            l = l * (1.0 + tint)
        else:
            l = l * (1.0 - tint) + HLSMAX * tint

        l = max(0.0, min(HLSMAX, l))

        r2, g2, b2 = colorsys.hls_to_rgb(
            h / HLSMAX,
            l / HLSMAX,
            s / HLSMAX,
        )

        return "{:02X}{:02X}{:02X}".format(
            round(r2 * 255),
            round(g2 * 255),
            round(b2 * 255),
        )
    except Exception:
        return rgb_hex.upper()


def color_to_hex(color, fallback=None, theme_colors=None):
    if color is None:
        return fallback

    try:
        typ = color.type

        if typ == "rgb" and color.rgb:
            s = str(color.rgb)[-6:]
            return f"#{s.upper()}"

        if typ == "indexed" and color.indexed is not None:
            idx = int(color.indexed)
            if 0 <= idx < len(INDEXED_COLORS):
                return "#" + INDEXED_COLORS[idx][-6:].upper()

        if typ == "theme" and color.theme is not None and theme_colors:
            idx = int(color.theme)
            if 0 <= idx < len(theme_colors):
                base = theme_colors[idx]
                if base:
                    resolved = _apply_excel_tint(base, getattr(color, "tint", 0.0))
                    if resolved:
                        return "#" + resolved

    except Exception:
        pass

    return fallback


def width_to_px(width):
    if width is None:
        width = 8.43
    return max(2, round(float(width) * 7 + 5, 1))

def row_height_to_px(height):
    if height is None:
        height = 15
    return max(2, round(float(height) * 96 / 72, 1))

def border_css(side, name, theme_colors=None):
    if side is None or side.style is None:
        return ""

    width = {
        "hair": "1px",
        "thin": "1px",
        "medium": "2px",
        "thick": "3px",
        "double": "3px",
    }.get(side.style, "1px")

    style_text = str(side.style)
    if side.style == "double":
        css_style = "double"
    elif "dash" in style_text:
        css_style = "dashed"
    elif "dot" in style_text:
        css_style = "dotted"
    else:
        css_style = "solid"

    color = color_to_hex(side.color, "#808080", theme_colors)
    return f"border-{name}:{width} {css_style} {color};"

def alignment_css(a):
    parts = []

    if a:
        h = a.horizontal
        v = a.vertical

        if h in ("center", "centerContinuous"):
            parts.append("text-align:center;")
        elif h == "right":
            parts.append("text-align:right;")
        elif h in ("justify", "distributed"):
            parts.append("text-align:justify;")
        else:
            parts.append("text-align:left;")

        if v == "top":
            parts.append("vertical-align:top;")
        elif v == "bottom":
            parts.append("vertical-align:bottom;")
        else:
            parts.append("vertical-align:middle;")

        if a.wrap_text:
            parts.append("white-space:normal;overflow-wrap:break-word;")
        else:
            parts.append("white-space:nowrap;")

    return "".join(parts)

def cell_css(cell, theme_colors=None):
    parts = []

    f = cell.font
    if f:
        if f.name:
            safe = f.name.replace("'", "\\'")
            parts.append(f"font-family:'{safe}',Arial,sans-serif;")
        if f.sz:
            parts.append(f"font-size:{float(f.sz):.2f}pt;")
        if f.bold:
            parts.append("font-weight:700;")
        if f.italic:
            parts.append("font-style:italic;")
        if f.underline:
            parts.append("text-decoration:underline;")
        c = color_to_hex(f.color, None, theme_colors)
        if c:
            parts.append(f"color:{c};")

    fill = cell.fill
    if fill and fill.fill_type:
        c = color_to_hex(fill.fgColor, None, theme_colors)
        if c:
            parts.append(f"background-color:{c};")

    parts.append(alignment_css(cell.alignment))

    b = cell.border
    if b:
        parts.append(border_css(b.left, "left", theme_colors))
        parts.append(border_css(b.right, "right", theme_colors))
        parts.append(border_css(b.top, "top", theme_colors))
        parts.append(border_css(b.bottom, "bottom", theme_colors))

    return "".join(parts)

def display_value(value):
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, time):
        return value.isoformat(timespec="seconds")

    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer():
            return str(int(value))
        return str(value)

    return str(value)

def build_html(src: Path, out: Path):
    wb = load_workbook(
        src,
        read_only=False,
        data_only=False,
        keep_vba=(src.suffix.lower() == ".xlsm"),
    )

    wb_values = load_workbook(
        src,
        read_only=False,
        data_only=True,
        keep_vba=False,
    )

    # Resolve workbook Theme Colors so ordinary Excel fill colors such as
    # Accent/Theme colors with tint can be reproduced in the HTML output.
    theme_colors = _theme_colors_from_workbook(wb)

    sheets = [name for name in WANTED_SHEETS if name in wb.sheetnames]

    if not sheets:
        sheets = [
            ws.title
            for ws in wb.worksheets
            if "Backup" not in ws.title
        ]

    chunks = []

    chunks.append("""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title}</title>
<style>
*{{box-sizing:border-box}}
html,body{{margin:0;padding:0;background:#f4f5f7;color:#111;font-family:Arial,sans-serif}}
.viewer-shell{{height:100vh;display:flex;flex-direction:column;min-width:0}}
.topbar{{flex:0 0 auto;background:#fff;border-bottom:1px solid #d7dbe0;padding:10px 12px 0}}
.book-title{{font-size:14px;font-weight:700;margin:0 0 8px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.tabs{{display:flex;gap:4px;overflow-x:auto}}
.tab{{border:1px solid #c9ced6;border-bottom:none;background:#eceff3;padding:8px 14px;cursor:pointer;border-radius:6px 6px 0 0;font-size:13px;white-space:nowrap}}
.tab.active{{background:#fff;font-weight:700}}
.sheet{{display:none;flex:1 1 auto;min-height:0;overflow:auto;background:#fff}}
.sheet.active{{display:block}}
.excel-table{{border-collapse:collapse;table-layout:fixed;width:max-content;min-width:max-content}}
.excel-table td{{padding:0 3px;box-sizing:border-box;overflow:visible}}
.notice{{padding:20px;font-size:14px}}
</style>
</head>
<body>
<div class="viewer-shell">
<div class="topbar">
<div class="book-title">{title}</div>
<div class="tabs">
""".format(title=html.escape(src.stem)))

    for i, name in enumerate(sheets):
        cls = "tab active" if i == 0 else "tab"
        chunks.append(
            f'<button class="{cls}" data-sheet="sheet{i}">'
            f'{html.escape(name)}</button>\n'
        )

    chunks.append("</div></div>\n")

    for i, name in enumerate(sheets):
        ws = wb[name]
        wsv = wb_values[name]

        cls = "sheet active" if i == 0 else "sheet"
        chunks.append(f'<section id="sheet{i}" class="{cls}">\n')

        visible_cols = [
            c
            for c in range(1, ws.max_column + 1)
            if not ws.column_dimensions[get_column_letter(c)].hidden
        ]

        visible_rows = [
            r
            for r in range(1, ws.max_row + 1)
            if not ws.row_dimensions[r].hidden
        ]

        if not visible_cols or not visible_rows:
            chunks.append(
                '<div class="notice">No visible cells.</div></section>\n'
            )
            continue

        chunks.append('<table class="excel-table"><colgroup>\n')

        for c in visible_cols:
            letter = get_column_letter(c)
            width = width_to_px(ws.column_dimensions[letter].width)
            chunks.append(f'<col style="width:{width}px">\n')

        chunks.append('</colgroup><tbody>\n')

        merged_lookup = {}

        for mr in ws.merged_cells.ranges:
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    merged_lookup[(r, c)] = mr

        emitted_merges = set()

        for r in visible_rows:
            row_px = row_height_to_px(ws.row_dimensions[r].height)
            chunks.append(f'<tr style="height:{row_px}px">\n')

            for c in visible_cols:
                mr = merged_lookup.get((r, c))

                if mr is not None:
                    key = str(mr)

                    if key in emitted_merges:
                        continue

                    vis_r = [
                        x for x in visible_rows
                        if mr.min_row <= x <= mr.max_row
                    ]

                    vis_c = [
                        x for x in visible_cols
                        if mr.min_col <= x <= mr.max_col
                    ]

                    if (
                        not vis_r
                        or not vis_c
                        or r != vis_r[0]
                        or c != vis_c[0]
                    ):
                        continue

                    src_cell = ws.cell(mr.min_row, mr.min_col)
                    val_cell = wsv.cell(mr.min_row, mr.min_col)

                    val = (
                        val_cell.value
                        if val_cell.value is not None
                        else src_cell.value
                    )

                    attrs = ""

                    if len(vis_c) > 1:
                        attrs += f' colspan="{len(vis_c)}"'

                    if len(vis_r) > 1:
                        attrs += f' rowspan="{len(vis_r)}"'

                    chunks.append(
                        f'<td{attrs} style="{cell_css(src_cell, theme_colors)}">'
                        f'{html.escape(display_value(val))}</td>\n'
                    )

                    emitted_merges.add(key)

                else:
                    src_cell = ws.cell(r, c)
                    val_cell = wsv.cell(r, c)

                    val = (
                        val_cell.value
                        if val_cell.value is not None
                        else src_cell.value
                    )

                    chunks.append(
                        f'<td style="{cell_css(src_cell, theme_colors)}">'
                        f'{html.escape(display_value(val))}</td>\n'
                    )

            chunks.append("</tr>\n")

        chunks.append("</tbody></table></section>\n")

    chunks.append("""
<script>
const tabs=[...document.querySelectorAll('.tab')];
const sheets=[...document.querySelectorAll('.sheet')];

tabs.forEach(btn=>{
  btn.addEventListener('click',()=>{
    tabs.forEach(x=>x.classList.remove('active'));
    sheets.forEach(x=>x.classList.remove('active'));
    btn.classList.add('active');
    document.getElementById(btn.dataset.sheet).classList.add('active');
  });
});
</script>
</div>
</body>
</html>
""")

    out.write_text("".join(chunks), encoding="utf-8")

    wb.close()
    wb_values.close()

def main():
    if len(sys.argv) > 1:
        selected = sys.argv[1]
    else:
        selected = pick_file()

    if not selected:
        return 0

    src = Path(selected).resolve()

    if not src.exists():
        show_error(
            "OST HTML Exporter",
            f"File not found:\n{src}"
        )
        return 1

    out = src.with_name(src.stem + "_Viewer.html")

    try:
        build_html(src, out)
    except Exception as exc:
        show_error(
            "OST HTML Exporter",
            f"Export failed:\n{type(exc).__name__}: {exc}"
        )
        return 1

    show_info(
        "OST HTML Exporter",
        f"Export completed:\n{out}"
    )

    try:
        import webbrowser
        webbrowser.open(out.as_uri())
    except Exception:
        pass

    return 0

if __name__ == "__main__":
    raise SystemExit(main())
